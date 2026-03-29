# app.py
import streamlit as st
import pandas as pd
import re
import requests
import tempfile
import os
import time
import openpyxl
from openpyxl import Workbook
from io import BytesIO
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from streamlit_pdf_viewer import pdf_viewer
from google import genai

# -----------------------
# CONFIG
# -----------------------
API_URL = "https://chat.binghamton.edu/api/chat/completions"
API_KEY = "" 
GEMINI_API_KEY = ""
gemini_client = genai.Client(api_key=GEMINI_API_KEY)

st.set_page_config(page_title="AEO Dashboard", layout="wide")

# -----------------------
# TABS
# -----------------------
tab1, tab2, tab3 = st.tabs(["AI Visibility & Sources", "Summary & Sources Report", "📊 Weekly Report"])

# -----------------------
# TAB 1: AI Visibility & Sources Generator
# -----------------------
with tab1:
    st.header("AI Visibility & Sources Generator")

    st.subheader("Models and Prompts")
    available_models = [
        "gpt-oss:120b",
        "codellama:70B",
        "gemma3:latest",
        "hermes3:latest",
        "llama3.1:70B",
        "mixtral:8x22b",
        "phi4:latest",
        "qwen2.5-coder:32B",
        "qwq:latest",
        "gemini-2.5-flash"
    ]
    col1, col2, col3, col4 = st.columns(4)
    models = []
    with col1:
        for model in available_models[:5]:
            if st.checkbox(model):
                models.append(model)
    
    with col2:
        for model in available_models[5:10]:
            if st.checkbox(model):
                models.append(model)

    prompts_str = st.text_area(
        "Prompts (one per line)",
        value=(
            "Best graduate schools in New York for computer science\n"
            "Top MS programs in engineering in the northeast USA\n"
            "Best public universities for master's degrees in New York\n"
            "Best ROI graduate programs in the USA"
        )
    )
    prompts = [p.strip() for p in prompts_str.split("\n") if p.strip()]

    output_file_name = "ollama_ai_visibility_results_dashboard.xlsx"

    # -------- Functions --------
    def query_model(model, prompt):

        prompt_with_instruction = (
            f"{prompt}\n\nPlease provide your answer in detail and include a section titled 'Sources' "
            "listing the main references or links you used."
        )

        # -----------------------
        # GEMINI ROUTE
        # -----------------------
        if "gemini" in model:

            try:
                response = gemini_client.models.generate_content(
                    model=model,
                    contents=prompt_with_instruction,
                    config={
                        "tools": [{"google_search": {}}]  
                    }
                )

                text_response = response.text

                metadata = response.candidates[0].grounding_metadata
                if metadata and metadata.web_search_queries:
                    searches = ", ".join(metadata.web_search_queries)
                    text_response += f"\n\n[Gemini Google Searches: {searches}]"

                return text_response

            except Exception as e:
                return f"GEMINI ERROR: {e}"


        # -----------------------
        # DEFAULT LLM API
        # -----------------------
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }

        data = {
            "model": model,
            "messages": [{"role": "user", "content": prompt_with_instruction}]
        }

        try:
            response = requests.post(API_URL, headers=headers, json=data, timeout=300)
            response.raise_for_status()
            json_resp = response.json()

            if "choices" in json_resp and json_resp["choices"]:
                return json_resp["choices"][0]["message"]["content"]

            return str(json_resp)

        except Exception as e:
            return f"ERROR: {e}"

    def extract_sources(text):
        urls = re.findall(r'(https?://[^\s)]+)', text)
        if urls:
            return "\n".join(urls)
        match = re.search(r'(Sources|References)[:\-]?\s*(.*)', text, re.IGNORECASE | re.DOTALL)
        if match:
            return match.group(2).strip()
        return "N/A"

    def generate_excel(models, prompts):
        if os.path.exists(output_file_name):
            wb = openpyxl.load_workbook(output_file_name)
        else:
            wb = Workbook()
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        for model in models:
            sheet_name = model[:6]
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Prompt", "Response", "Sources"])

            for prompt in prompts:
                st.info(f"Querying model {model} for prompt: {prompt}")
                answer = query_model(model, prompt)
                sources = extract_sources(answer)
                ws.append([prompt, answer, sources])
                st.success(f"Completed prompt → waiting 15 seconds...")
                time.sleep(15)

        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)
        return excel_bytes

    # -------- Generate Button --------
    if st.button("Generate Excel"):
        if not models or not prompts:
            st.error("Please provide at least one model and one prompt.")
        else:
            st.info("Generating Excel, this may take several minutes...")
            excel_data = generate_excel(models, prompts)
            st.success("Excel generation complete!")
            st.download_button(
                "📥 Download Excel File",
                data=excel_data,
                file_name=output_file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# -----------------------
# TAB 2: Summary & Sources Report
# -----------------------
with tab2:
    st.header("Summary & Sources Report")

    st.subheader("Upload Excel & Generate Reports")
    uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])
    options = [
        "mixtral:8x22b",
        "gpt-oss:120b",
        "codellama:70B",
        "gemma3:latest",
        "hermes3:latest",
        "llama3.1:70B",
        "phi4:latest",
        "qwen2.5-coder:32B",
        "qwq:latest",
    ]

    model_name = st.selectbox("LLM Model Name", options)

    summary_btn = st.button("Generate Summary Report")
    sources_btn = st.button("Generate Sources Report")

    # -------- UTILITY FUNCTIONS --------
    def chunk_list(lst, chunk_size):
        for i in range(0, len(lst), chunk_size):
            yield lst[i:i + chunk_size]

    def safe_ollama_chat(model, prompt):
        headers = {"Content-Type": "application/json", "Authorization": f"Bearer {API_KEY}"}
        payload = {"model": model, "messages": [{"role": "user", "content": prompt}]}

        try:
            response = requests.post(API_URL, headers=headers, json=payload, timeout=300)
            response.raise_for_status()

            data = response.json()
            return data["choices"][0]["message"]["content"]

        except Exception as e:
            return f"LLM ERROR: {str(e)}"
    
    # -----------------------
# SUMMARY REPORT
# -----------------------
    def generate_aeo_report_pdf(file_path, model):
        xls = pd.ExcelFile(file_path)
        all_rows = []

        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            for _, row in df.iterrows():
                all_rows.append(f"Prompt: {row['Prompt']}\nResponse: {row['Response']}")

        def chunk_list(lst, chunk_size):
            for i in range(0, len(lst), chunk_size):
                yield lst[i:i + chunk_size]

        chunks = list(chunk_list(all_rows, 100))
        chunk_summaries = []

        for idx, chunk in enumerate(chunks):
            chunk_text = "\n\n".join(chunk)
            chunk_prompt = f"""
    You are analyzing AI-generated responses about graduate program flexibility.
    Summarize ONLY the following dataset chunk into 6-10 factual bullet points:

    {chunk_text}
    """
            summary = safe_ollama_chat(model, chunk_prompt)
            chunk_summaries.append(summary)

        # Final synthesis
        final_prompt = f"""
    You are an expert AEO (AI Engine Optimization) and graduate program analyst.
    You are given summaries of AI model responses about graduate program flexibility.

    Your job is to analyze ONLY this provided data.

    Create a structured executive report with the following sections:

    1. Executive Summary
    - 3-5 concise bullet points capturing the key findings.

    2. University Mention Frequency
    - Identify which universities appear the most.
    - Rank them by number of flexibility-related mentions.
    - Provide a short ranked list.

    3. Top Programs Offering Flexible Work Arrangements
    - Extract programs explicitly described as flexible.
    - Identify the type of flexibility mentioned.

    4. Binghamton University Coverage
    - How often Binghamton appears.
    - What is said about flexibility.
    - Where competitors appear but Binghamton does not.

    5. Competitor Gap Analysis
    - Identify universities frequently mentioned where Binghamton is not.
    - Summarize what competitors emphasize.

    6. Actionable Recommendations for Binghamton University
    - Provide 5-7 data-driven recommendations tied directly to gaps in the dataset.

    Format the output with clear headers and bullet points.

    Here are the chunk summaries:
    {chr(10).join(chunk_summaries)}
    """
        final_report = safe_ollama_chat(model, final_prompt)
        final_report_clean = re.sub(r"\n{3,}", "\n\n", final_report)

        # Build PDF
        output_file = file_path.replace(".xlsx", "_LLM_Report_Final_dashboard.pdf")
        doc = SimpleDocTemplate(output_file, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        for line in final_report_clean.split("\n"):
            if line.strip():
                story.append(Paragraph(line.strip(), styles["Normal"]))
                story.append(Spacer(1, 10))

        chart_paths = generate_aeo_charts(file_path)
        story.append(Spacer(1, 20))
        story.append(Paragraph("<b>Visual Analytics</b>", styles["Heading2"]))
        story.append(Spacer(1, 10))

        for chart_path in chart_paths:
            story.append(Image(chart_path, width=500, height=350))
            story.append(Spacer(1, 20))

        doc.build(story)
        return output_file

    # -----------------------
    # SOURCES REPORT
    # -----------------------
    def classify_source(domain):
        d = domain.lower()
        if any(x in d for x in ["facebook", "linkedin", "twitter", "instagram", "youtube", "tiktok", "reddit"]):
            return "Social Media"
        elif ".edu" in d:
            return "University Website"
        elif ".gov" in d or ".ac" in d:
            return "Government/Educational Body"
        elif any(x in d for x in [".com", ".org", ".net", "news", "college", "ranking", "review"]):
            return "News / Articles / Blogs"
        else:
            return "Miscellaneous"

    def extract_all_urls(text):
        if not isinstance(text, str): return []
        text = text.replace(")", " ").replace("]", " ")
        urls = re.findall(r'https?://[A-Za-z0-9\.\-_/~%?#=&]+', text)
        clean_domains = []
        for u in urls:
            u = u.strip().strip(".,);:]")
            m = re.findall(r'https?://(?:www\.)?([^/]+)/?', u)
            clean_domains.append(m[0].lower() if m else u)
        return list(set(clean_domains))

    def generate_sources_report_pdf(file_path, model):
        xls = pd.ExcelFile(file_path)
        all_sources = []

        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df.columns = [c.lower() for c in df.columns]

            src_col = next((c for c in df.columns if "source" in c), None)
            prompt_col = next((c for c in df.columns if "prompt" in c), None)
            if not src_col or not prompt_col: continue

            for _, row in df.iterrows():
                urls = extract_all_urls(str(row[src_col]))
                for d in urls:
                    all_sources.append({
                        "Model": sheet,
                        "Prompt": row[prompt_col],
                        "Website": d,
                        "Category": classify_source(d)
                    })

        df_sources = pd.DataFrame(all_sources)
        if df_sources.empty: return None

        agg = df_sources.groupby(["Category", "Website"]).size().reset_index(name="Count").sort_values(["Category", "Count"], ascending=[True, False])

        # Pie chart
        category_counts = df_sources["Category"].value_counts()
        plt.figure(figsize=(7, 7))
        plt.pie(category_counts, labels=category_counts.index, autopct="%1.1f%%")
        chart_path = os.path.join(tempfile.gettempdir(), "source_pie_chart.png")
        plt.savefig(chart_path, dpi=300, bbox_inches="tight")
        plt.close()

        # LLM summary
        summary_prompt = f"""
    You are analyzing sources used by multiple AI models for graduate program prompts.

    Models: {', '.join(sorted(df_sources['Model'].unique()))}
    Total sources: {len(df_sources)}
    Category breakdown: {category_counts.to_dict()}
    Top 10 websites: {df_sources['Website'].value_counts().head(10).to_dict()}

    Write a short bullet-point summary describing:
    - General source patterns
    - Any biases or missing source types
    """
        llm_summary = safe_ollama_chat(model, summary_prompt)

        # Build PDF
        out_pdf = file_path.replace(".xlsx", "_Sources_Report_dashboard.pdf")
        doc = SimpleDocTemplate(out_pdf, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph("<b>1. Summary</b>", styles["Heading2"]))
        story.append(Spacer(1, 10))
        for line in llm_summary.split("\n"):
            if line.strip():
                story.append(Paragraph(line.strip(), styles["Normal"]))
                story.append(Spacer(1, 6))

        story.append(Spacer(1, 15))
        story.append(Paragraph("<b>2. Source Type Distribution</b>", styles["Heading2"]))
        story.append(Spacer(1, 10))
        story.append(Image(chart_path, width=400, height=400))
        story.append(Spacer(1, 20))

        story.append(Paragraph("<b>3. Source Table</b>", styles["Heading2"]))
        story.append(Spacer(1, 8))

        for cat in agg["Category"].unique():
            subset = agg[agg["Category"] == cat]
            story.append(Paragraph(f"<b>{cat}</b>", styles["Heading3"]))
            data = [["Website", "Count"]] + subset[["Website", "Count"]].values.tolist()
            t = Table(data, repeatRows=1, colWidths=[300, 80])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]))
            story.append(t)
            story.append(Spacer(1, 14))

        doc.build(story)
        return out_pdf


    def generate_aeo_charts(file_path):
        universities = ["Binghamton", "Buffalo", "Stony Brook", "Columbia", "NYU", "Cornell",
                        "Syracuse", "RIT", "SUNY", "CUNY", "RPI", "University at Albany"]
        xls = pd.ExcelFile(file_path)
        all_data = []
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            for _, row in df.iterrows():
                response = str(row["Response"])
                prompt = row["Prompt"]
                for uni in universities:
                    count = len(re.findall(rf"\b{re.escape(uni)}\b", response, re.IGNORECASE))
                    all_data.append({"Model": sheet, "Prompt": prompt, "University": uni, "Mentions": count})
        df_all = pd.DataFrame(all_data)
        chart_paths = []

        mentions_sum = df_all.groupby("University")["Mentions"].sum().sort_values(ascending=False)
        plt.figure(figsize=(10, 6))
        mentions_sum.plot(kind="bar")
        plt.title("Total Mentions per University (All Models)")
        plt.ylabel("Number of Mentions")
        plt.xlabel("University")
        plt.xticks(rotation=45)
        plt.tight_layout()
        bar_chart_path = os.path.join(tempfile.gettempdir(), "mentions_bar_chart.png")
        plt.savefig(bar_chart_path)
        chart_paths.append(bar_chart_path)
        plt.close()

        pivot_prompt_uni = df_all.groupby(["Prompt", "University"])["Mentions"].sum().unstack().fillna(0)
        plt.figure(figsize=(12, 10))
        sns.heatmap(pivot_prompt_uni, cmap="YlGnBu", linewidths=0.5)
        plt.title("Heatmap: Prompts vs University Mentions")
        plt.xlabel("University")
        plt.ylabel("Prompt")
        plt.tight_layout()
        heatmap_path = os.path.join(tempfile.gettempdir(), "mentions_heatmap.png")
        plt.savefig(heatmap_path)
        chart_paths.append(heatmap_path)
        plt.close()
        return chart_paths

    if uploaded_file:
        temp_path = "uploaded.xlsx"
        with open(temp_path, "wb") as f:
            f.write(uploaded_file.getbuffer())

        if summary_btn:
            st.subheader("Summary Report")
            st.info("Generating... please wait.")
            pdf_path = generate_aeo_report_pdf(temp_path, model_name)
            with open(pdf_path, "rb") as f:
                pdf_bytes = f.read()
            st.download_button("📥 Download Summary PDF", data=pdf_bytes,
                               file_name=os.path.basename(pdf_path), mime="application/pdf")
            st.subheader("📄 PDF Preview")
            pdf_viewer(pdf_bytes)

        if sources_btn:
            st.subheader("Sources Report")
            st.info("Generating... please wait.")
            pdf_path = generate_sources_report_pdf(temp_path, model_name)
            if pdf_path:
                with open(pdf_path, "rb") as f:
                    pdf_bytes = f.read()
                st.download_button("📥 Download Sources PDF", data=pdf_bytes,
                                   file_name=os.path.basename(pdf_path), mime="application/pdf")
                st.subheader("📄 PDF Preview")
                pdf_viewer(pdf_bytes)
            else:
                st.error("No valid sources found in the file.")
    else:
        st.info("Upload an Excel file to begin.")
# -----------------------
# TAB 3: Weekly AEO Report (Conductor-style)
# -----------------------
with tab3: 
    st.header("📊 Weekly AEO Report")
    st.caption("Conductor Monitoring-style weekly digest of your AI visibility data")

    uploaded_weekly = st.file_uploader("Upload This Week's Excel", type=["xlsx"], key="weekly_current")
    uploaded_prev = st.file_uploader("Upload Last Week's Excel (optional, for delta)", type=["xlsx"], key="weekly_prev")

    target_university = st.text_input("Your University (for Health scoring)", value="Binghamton")
    competitors = st.text_input(
        "Competitor Universities (comma-separated)",
        value="Buffalo, Stony Brook, Columbia, NYU, Cornell, Syracuse, RIT"
    )

    weekly_btn = st.button("Generate Weekly Report")

    # -------- Helper: Compute metrics from Excel --------
    def compute_weekly_metrics(file_bytes, target, comp_list):
        xls = pd.ExcelFile(file_bytes)
        rows = []
        for sheet in xls.sheet_names:
            df = pd.read_excel(xls, sheet_name=sheet)
            df.columns = [c.lower() for c in df.columns]
            for _, row in df.iterrows():
                response = str(row.get("response", ""))
                prompt = str(row.get("prompt", ""))
                target_count = len(re.findall(rf"\b{re.escape(target)}\b", response, re.IGNORECASE))
                comp_counts = sum(
                    len(re.findall(rf"\b{re.escape(c.strip())}\b", response, re.IGNORECASE))
                    for c in comp_list
                )
                urls = extract_all_urls(str(row.get("sources", "")))
                rows.append({
                    "Model": sheet,
                    "Prompt": prompt,
                    "Target_Mentions": target_count,
                    "Competitor_Mentions": comp_counts,
                    "Source_Count": len(urls),
                    "Has_Target": target_count > 0,
                })
        return pd.DataFrame(rows)

    def compute_health_score(df):
        """Score 0-100: % of responses where target university appears"""
        if df.empty: return 0
        return round((df["Has_Target"].sum() / len(df)) * 100, 1)

    def compute_issues(df, comp_list):
        """Issues = prompts where competitors appear but target does not"""
        issues = df[(df["Has_Target"] == False) & (df["Competitor_Mentions"] > 0)].sort_values("Competitor_Mentions", ascending=False).drop_duplicates(subset=["Prompt"])
        return issues

    def build_conductor_pdf(metrics_now, metrics_prev, health_now, health_prev,
                             issues_now, issues_prev, target, output_path):
        doc = SimpleDocTemplate(output_path, pagesize=letter,
                                rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        story = []

        # Title
        from reportlab.lib.styles import ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        title_style = ParagraphStyle("title", fontSize=18, fontName="Helvetica-Bold",
                                     spaceAfter=4, alignment=TA_CENTER)
        sub_style = ParagraphStyle("sub", fontSize=10, textColor=colors.grey,
                                   spaceAfter=20, alignment=TA_CENTER)
        story.append(Paragraph("Weekly AEO Monitoring Report", title_style))
        story.append(Paragraph(f"Week ending {pd.Timestamp.now().strftime('%B %d, %Y')} · Target: {target}", sub_style))

        # ---- Summary row (4 KPI boxes) ----
        health_delta = round(health_now - health_prev, 1) if health_prev is not None else None
        delta_str = (f"{'▲' if health_delta >= 0 else '▼'} {abs(health_delta)}" if health_delta is not None else "—")
        delta_color = colors.HexColor("#27ae60") if (health_delta or 0) >= 0 else colors.HexColor("#e74c3c")

        total_pages = len(metrics_now)
        prev_prompts = set(metrics_prev["Prompt"]) if metrics_prev is not None else set()
        curr_prompts = set(metrics_now["Prompt"])
        added = len(curr_prompts - prev_prompts)
        removed = len(prev_prompts - curr_prompts)
        changed = len(curr_prompts & prev_prompts)  # treated as "checked again"

        issues_open = len(issues_now)
        issues_prev_count = len(issues_prev) if issues_prev is not None else issues_open
        issues_resolved = max(0, issues_prev_count - issues_open)
        issues_opened = max(0, issues_open - issues_prev_count)

        # KPI table
        kpi_data = [
            [
                Paragraph("<b>SCOPE</b>", styles["Normal"]),
                Paragraph("<b>HEALTH</b>", styles["Normal"]),
                Paragraph("<b>PAGES / PROMPTS</b>", styles["Normal"]),
                Paragraph("<b>ISSUES</b>", styles["Normal"]),
            ],
            [
                Paragraph(f"All Models<br/><font size=8 color='grey'>{len(metrics_now['Model'].unique())} model(s)</font>", styles["Normal"]),
                Paragraph(f"<font size=24><b>{health_now}%</b></font><br/><font color='{'green' if (health_delta or 0) >= 0 else 'red'}'>{delta_str} vs last week</font>", styles["Normal"]),
                Paragraph(f"<font size=24><b>{total_pages}</b></font><br/>"
                          f"<font color='orange'>~ {changed} checked</font>  "
                          f"<font color='green'>+{added} new</font>  "
                          f"<font color='red'>-{removed} removed</font>", styles["Normal"]),
                Paragraph(f"<font size=24><b>{issues_open}</b></font><br/>"
                          f"<font color='red'>+{issues_opened} opened</font>  "
                          f"<font color='green'>✓{issues_resolved} resolved</font>", styles["Normal"]),
            ]
        ]

        kpi_table = Table(kpi_data, colWidths=[110, 120, 180, 120])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor("#ecf0f1")),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor("#ecf0f1"), colors.white]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#bdc3c7")),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 20))

        # ---- Per-model breakdown ----
        story.append(Paragraph("<b>Model-by-Model Breakdown</b>", styles["Heading2"]))
        story.append(Spacer(1, 8))

        model_rows = [["Model", "Health %", "Prompts", f"{target} Mentions", "Avg Sources"]]
        for model_id in sorted(metrics_now["Model"].unique()):
            sub = metrics_now[metrics_now["Model"] == model_id]
            mh = round((sub["Has_Target"].sum() / len(sub)) * 100, 1)
            model_rows.append([
                model_id,
                f"{mh}%",
                len(sub),
                int(sub["Target_Mentions"].sum()),
                round(sub["Source_Count"].mean(), 1),
            ])

        model_table = Table(model_rows, repeatRows=1, colWidths=[100, 70, 70, 100, 80])
        model_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f6fc")]),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(model_table)
        story.append(Spacer(1, 20))

        # ---- Issues table ----
        story.append(Paragraph(f"<b>Open Issues — Prompts where {target} is missing but competitors appear</b>", styles["Heading2"]))
        story.append(Spacer(1, 8))

        if issues_now.empty:
            story.append(Paragraph("✅ No open issues this week!", styles["Normal"]))
        else:
            issue_rows = [["Model", "Prompt", "Competitor Mentions"]]
            for _, r in issues_now.iterrows():
                prompt_short = r["Prompt"][:80] + "..." if len(r["Prompt"]) > 80 else r["Prompt"]
                issue_rows.append([r["Model"], prompt_short, int(r["Competitor_Mentions"])])

            issue_table = Table(issue_rows, repeatRows=1, colWidths=[80, 310, 100])
            issue_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#c0392b")),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor("#fdedec")]),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('WORDWRAP', (1, 0), (1, -1), True),
            ]))
            story.append(issue_table)

        story.append(Spacer(1, 20))

        # ---- Bar chart: Health by model ----
        models_list = sorted(metrics_now["Model"].unique())
        health_vals = [
            round((metrics_now[metrics_now["Model"] == m]["Has_Target"].sum() /
                   len(metrics_now[metrics_now["Model"] == m])) * 100, 1)
            for m in models_list
        ]
        fig, ax = plt.subplots(figsize=(8, 4))
        bar_colors = ["#27ae60" if v >= 50 else "#e74c3c" for v in health_vals]
        bars = ax.bar(models_list, health_vals, color=bar_colors)
        ax.axhline(y=50, color="orange", linestyle="--", linewidth=1, label="50% threshold")
        ax.set_title(f"Health Score by Model (% responses mentioning {target})")
        ax.set_ylabel("Health %")
        ax.set_ylim(0, 105)
        ax.set_xticklabels(models_list, rotation=30, ha="right")
        for bar, val in zip(bars, health_vals):
            ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1, f"{val}%", ha="center", fontsize=8)
        plt.tight_layout()
        chart_buf = BytesIO()
        plt.savefig(chart_buf, format="png", dpi=150)
        chart_buf.seek(0)
        plt.close()

        from reportlab.lib.utils import ImageReader
        story.append(Paragraph("<b>Health Score by Model</b>", styles["Heading2"]))
        story.append(Spacer(1, 8))
        story.append(Image(chart_buf, width=480, height=240))

        doc.build(story)

    # ---- Run on button click ----
    if weekly_btn:
        if not uploaded_weekly:
            st.error("Please upload this week's Excel file.")
        else:
            comp_list = [c.strip() for c in competitors.split(",") if c.strip()]

            metrics_now = compute_weekly_metrics(uploaded_weekly, target_university, comp_list)
            health_now = compute_health_score(metrics_now)
            issues_now = compute_issues(metrics_now, comp_list)

            metrics_prev = None
            health_prev = None
            issues_prev = None
            if uploaded_prev:
                metrics_prev = compute_weekly_metrics(uploaded_prev, target_university, comp_list)
                health_prev = compute_health_score(metrics_prev)
                issues_prev = compute_issues(metrics_prev, comp_list)

            # ---- Streamlit live preview (Conductor-style cards) ----
            health_delta = round(health_now - health_prev, 1) if health_prev is not None else None
            prev_prompts = set(metrics_prev["Prompt"]) if metrics_prev is not None else set()
            curr_prompts = set(metrics_now["Prompt"])
            added = len(curr_prompts - prev_prompts)
            removed = len(prev_prompts - curr_prompts)
            issues_open = len(issues_now)
            issues_prev_count = len(issues_prev) if issues_prev is not None else issues_open

            st.markdown("---")
            st.subheader("📋 Weekly Snapshot")

            col_scope, col_health, col_pages, col_issues = st.columns(4)

            with col_scope:
                st.markdown("**🔭 Scope**")
                st.metric("Models", len(metrics_now["Model"].unique()))
                st.caption(f"{len(metrics_now)} total responses")

            with col_health:
                st.markdown("**💚 Health**")
                st.metric(
                    "Visibility Score",
                    f"{health_now}%",
                    delta=f"{health_delta:+.1f}%" if health_delta is not None else None
                )
                st.caption(f"% of responses mentioning {target_university}")

            with col_pages:
                st.markdown("**📄 Prompts / Pages**")
                st.metric("Total Prompts", len(curr_prompts))
                st.markdown(
                    f"🟠 Recurring: **{len(curr_prompts & prev_prompts)}** &nbsp;"
                    f"🟢 New: **+{added}** &nbsp;"
                    f"🔴 Dropped: **-{removed}**"
                )

            with col_issues:
                st.markdown("**⚠️ Issues**")
                issues_resolved = max(0, issues_prev_count - issues_open)
                issues_opened_new = max(0, issues_open - issues_prev_count)
                st.metric("Open Issues", issues_open,
                          delta=f"{issues_open - issues_prev_count:+d}" if issues_prev is not None else None,
                          delta_color="inverse")
                st.markdown(f"🔴 Opened: **{issues_opened_new}** &nbsp; 🟢 Resolved: **{issues_resolved}**")

            st.markdown("---")

            # Model breakdown table in UI
            st.subheader("Model Breakdown")
            breakdown_rows = []
            for m in sorted(metrics_now["Model"].unique()):
                sub = metrics_now[metrics_now["Model"] == m]
                mh = round((sub["Has_Target"].sum() / len(sub)) * 100, 1)
                breakdown_rows.append({
                    "Model": m,
                    "Health %": f"{mh}%",
                    "Prompts": len(sub),
                    f"{target_university} Mentions": int(sub["Target_Mentions"].sum()),
                    "Avg Sources/Response": round(sub["Source_Count"].mean(), 1),
                })
            st.dataframe(pd.DataFrame(breakdown_rows), use_container_width=True)

            if not issues_now.empty:
                st.subheader(f"⚠️ Prompts {target_university} Missed ({issues_open})")
                st.dataframe(issues_now[["Model", "Prompt", "Competitor_Mentions"]], use_container_width=True)
            else:
                st.success("✅ No open issues — your university appeared in all competitive responses!")

            # Generate PDF
            st.markdown("---")
            with st.spinner("Building PDF report..."):
                pdf_out = os.path.join(tempfile.gettempdir(), "weekly_aeo_report.pdf")
                build_conductor_pdf(
                    metrics_now, metrics_prev, health_now, health_prev,
                    issues_now, issues_prev, target_university, pdf_out
                )
                with open(pdf_out, "rb") as f:
                    pdf_bytes = f.read()

            st.download_button("📥 Download Weekly Report PDF", data=pdf_bytes,
                               file_name="weekly_aeo_report.pdf", mime="application/pdf")
            st.subheader("📄 PDF Preview")
            pdf_viewer(pdf_bytes)